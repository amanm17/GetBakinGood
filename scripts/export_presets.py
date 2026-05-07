#!/usr/bin/env python3
"""
Export presets & conversion rules from GetBakinGood.xlsx into JSON for the website.

Usage:
  python scripts/export_presets.py --xlsx data/GetBakinGood.xlsx --out data --public public/data
"""
import argparse, json, os
from datetime import datetime
from openpyxl import load_workbook

def read_sheet_rows(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        yield row

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="data/GetBakinGood.xlsx")
    ap.add_argument("--out", default="data")
    ap.add_argument("--public", default="public/data")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)

    presets = []
    ws = wb["Presets"]
    for row in read_sheet_rows(ws, 2):
        if not row[0]:
            continue
        presets.append({
            "dish_id": row[0],
            "category": row[1],
            "dish": row[2],
            "mode_type": row[3],
            "base_temp_c": row[4],
            "base_time_min": row[5],
            "base_preheat_min": row[6],
            "default_container": row[7],
            "default_rack": row[8],
            "base_qty": row[9],
            "notes": row[10],
        })

    rules = []
    ws = wb["Rules"]
    for row in read_sheet_rows(ws, 2):
        if not row[0]:
            continue
        rules.append({
            "key": row[0],
            "device": row[1],
            "mode_type": row[2],
            "temp_offset_c": row[3],
            "time_factor": row[4],
            "preheat_min": row[5],
            "notes": row[6],
        })

    meta = {
        "name": "GetBakinGood",
        "version": "1.0",
        "generated_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "units": {"temp": "C", "time": "min"},
        "source_xlsx": os.path.basename(args.xlsx),
    }

    os.makedirs(args.out, exist_ok=True)
    os.makedirs(args.public, exist_ok=True)

    presets_path = os.path.join(args.out, "presets.json")
    rules_path = os.path.join(args.out, "conversion_rules.json")
    with open(presets_path, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "presets": presets}, f, ensure_ascii=False, indent=2)
    with open(rules_path, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "rules": rules}, f, ensure_ascii=False, indent=2)

    # copy to public/data
    for src in (presets_path, rules_path):
        dst = os.path.join(args.public, os.path.basename(src))
        with open(src, "rb") as fsrc, open(dst, "wb") as fdst:
            fdst.write(fsrc.read())

    print("Exported:")
    print(" -", presets_path)
    print(" -", rules_path)
    print("Copied to:")
    print(" -", os.path.join(args.public, "presets.json"))
    print(" -", os.path.join(args.public, "conversion_rules.json"))

if __name__ == "__main__":
    main()
